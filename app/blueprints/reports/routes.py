from flask import Blueprint, jsonify, request, send_from_directory, current_app, send_file
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import io
import pytz  # If you need to handle timezones


reports_blueprint = Blueprint('reports', __name__)


@reports_blueprint.route('/customer-select', methods=['GET'])
def customer_select():
    bigquery_client = current_app.bigquery_client_UK
    query = f"""SELECT id,name FROM `machinemax-prod-635d.google_cloud_postgresql_public.organisations`"""

    try:
            query_job = bigquery_client.query(query)
            results = query_job.result()  # Waits for the job to complete.

            # Process the results
            data = [{"value": row.id,
                     "name": row.name,
                     } for row in results]

            print('done')

            return jsonify({'data':data})

    except Exception as e:
            print(e)
            return jsonify({"error": str(e)}), 500

def convert_to_naive_utc(dt):
    """Convert a timezone-aware datetime to a timezone-naive UTC datetime, or return None if dt is None."""
    if dt is None:
        return None
    if dt.tzinfo is not None:
        # Convert to UTC and make timezone naive
        return dt.astimezone(pytz.utc).replace(tzinfo=None)
    return dt


@reports_blueprint.route('/reports-data_completeness/<customer>', methods=['GET'])
def data_completeness(customer):
    bigquery_client = current_app.bigquery_client_UK

    query = f"""
WITH DistinctTableA AS (
    SELECT DISTINCT source_id, machine_id
    FROM `machinemax-prod-635d.google_cloud_postgresql_public.source_associations`
),
DistinctTableB AS (
    SELECT DISTINCT source_id, deveui
    FROM `machinemax-prod-635d.google_cloud_postgresql_public.latest_uplink`
)

SELECT
    DA.machine_id,
    M.asset_id,
    M.name,
    DT_A.source_id,
    FD.model,
    DT_B.deveui,
    
    COALESCE(MAX(LS.battery_level), MAX(LST.external_voltage)) AS battery_voltage,
    COALESCE(MAX(LS.updated_at), MAX(LST.updated_at)) AS last_updated_at,
    SUM(DA.off_hours + DA.total_hours) AS data_completeness,
    CASE
        WHEN DATE(COALESCE(MAX(LS.updated_at), MAX(LST.updated_at))) = CURRENT_DATE() THEN 'Responsive'
        ELSE 'Unresponsive'
    END AS status

FROM
    `machinemax-prod-635d.google_cloud_postgresql_public.daily_aggregate` DA
INNER JOIN 
    `machinemax-prod-635d.google_cloud_postgresql_public.machines` M on M.ID = DA.machine_id

LEFT JOIN
    DistinctTableA DT_A ON DA.machine_id = DT_A.machine_id
LEFT JOIN
    DistinctTableB DT_B ON DT_A.source_id = DT_B.source_id
LEFT JOIN 
    `machinemax-prod-635d.google_cloud_postgresql_public.latest_status` LS on LS.deveui = DT_B.deveui
LEFT JOIN
    `machinemax-prod-635d.google_cloud_postgresql_public.fotaweb_device` FD on CAST(FD.imei AS STRING) = DT_B.deveui
LEFT JOIN 
    `machinemax-prod-635d.google_cloud_postgresql_public.latest_status_teltonika` LST on LST.source_id = DT_B.source_id
WHERE
    DA.organisation_id = '{customer}'
    AND DA.date >= DATE_SUB(CURRENT_DATE, INTERVAL 2 DAY)
    AND DA.date <= CURRENT_DATE
    AND M.archived IS NULL
GROUP BY
    DA.machine_id, DT_A.source_id, DT_B.deveui, M.name, M.asset_id, FD.model;
"""



    try:
        query_job = bigquery_client.query(query)
        results = query_job.result()  # Waits for the job to complete.

        # Process the results
        # data = [{"machine_id": row.machine_id,
        #          "asset_id": row.asset_id,
        #          "source_id": row.source_id,
        #          "model": row.model,
        #          "deveui": row.deveui,
        #          "last_updated_at": row.last_updated_at,
        #          "data_completeness":row.data_completeness,
        #          "status":row.status,
        #          } for row in results]

        wb = Workbook()
        ws = wb.active

        # Add headers
        headers = ["IMEI","Model", "Machine ID", "Asset ID" , "MachineName","Status", "Battery Voltage", "Data Completeness"]
        ws.append(headers)

        def sort_key(row):
            # Custom sort function: 'Unresponsive' first, then 'Responsive', then others,
            # and within the same 'status', sort by 'model'
            status_order = {'Responsive': 0, 'Unresponsive': 1}
            status_value = status_order.get(row.status, 2)

            # If row.status is None, assign a value greater than 2 to ensure it comes after 'Unresponsive' and 'Responsive'
            if status_value == 2:
                return (status_value, row.model)
            else:
                return (status_value, row.model) if row.model else (status_value, "")

        sorted_results = sorted(results, key=sort_key)

        current_row = 2
        status_col_index = 6  # Adjust as per your Excel sheet
        battery_voltage_col_index = 7  # Adjust as per your Excel sheet
        data_completeness_col_index = 8  # Adjust as per your Excel sheet

        column_widths = {
            1: 20,  # Column A width
            2: 10,  # Column B width
            3: 40,  # Column C width
            4: 12,  # Column D width
            5: 30,  # Column E width
            6: 12,  # Column E width
            7: 14,  # Column E width
            8: 14,  # Column E width
        }

        # Set column widths based on the dictionary
        for col, width in column_widths.items():
            ws.column_dimensions[chr(64 + col)].width = width

        for row in sorted_results:
            # Define your row data
            last_updated_at_naive = convert_to_naive_utc(row.last_updated_at)

            row_data = [row.deveui, row.model, row.machine_id, row.asset_id, row.name, row.status, row.battery_voltage, row.data_completeness,]

            ws.append(row_data)

            # Apply coloring based on condition
            if row.status == 'Unresponsive':
                ws.cell(row=current_row, column=status_col_index).fill = PatternFill(start_color="FFC7CE",
                                                                                     end_color="FFC7CE",
                                                                                     fill_type="solid")
            elif row.status == 'Responsive':
                ws.cell(row=current_row, column=status_col_index).fill = PatternFill(start_color="C6EFCE",
                                                                                     end_color="C6EFCE",
                                                                                     fill_type="solid")

            try:
                if float(row.data_completeness) > 71.3:
                    ws.cell(row=current_row, column=data_completeness_col_index).fill = PatternFill(start_color="C6EFCE",
                                                                                                    end_color="C6EFCE",
                                                                                                    fill_type="solid")
                elif float(row.data_completeness) > 3.1:
                    ws.cell(row=current_row, column=data_completeness_col_index).fill = PatternFill(start_color="FFEB9C",
                                                                                                    end_color="FFEB9C",
                                                                                                    fill_type="solid")
                else:
                    ws.cell(row=current_row, column=data_completeness_col_index).fill = PatternFill(start_color="FFC7CE",
                                                                                                    end_color="FFC7CE",
                                                                                                    fill_type="solid")
            except:
                pass

            try:
                if float(row.battery_voltage) < 10000 and row.model == "FMT100":
                    ws.cell(row=current_row, column=battery_voltage_col_index).fill = PatternFill(
                        start_color="FFC7CE",
                        end_color="FFC7CE",
                        fill_type="solid")
                elif float(row.battery_voltage) < 2700 and row.model != "FMT100":
                    ws.cell(row=current_row, column=battery_voltage_col_index).fill = PatternFill(
                        start_color="FFC7CE",
                        end_color="FFC7CE",
                        fill_type="solid")
                else:
                    pass
            except:
                pass

            current_row += 1

        # Save the workbook to a BytesIO object
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Return the Excel file to download
        return send_file(
            excel_file,
            as_attachment=True,
            download_name="Report.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        # return jsonify({'report':data})

    except Exception as e:
        print(e)
        return jsonify({"error": str(e)}), 500
