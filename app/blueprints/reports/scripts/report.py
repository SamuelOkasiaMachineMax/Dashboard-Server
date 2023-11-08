from app import create_app  # Import your Flask app instance
from app.models import Customer, Machine, Sensor
from prettytable import PrettyTable
import pandas as pd
import openpyxl.styles


def show():
    headers = [
        'id',
        'asset_id',
        'machine_name',
        'source_status',
        'battery_voltage',
        'signal_status',
        'device_model',
        'customer_id'
    ]
    table = PrettyTable()
    table.field_names = headers
    for line in result:
        table.add_row([line[header] for header in headers])

    print(table)

def view():
    headers = sorted_df.columns.tolist()
    table = PrettyTable()
    table.field_names = headers

    for index, row in sorted_df.iterrows():
        table.add_row(row)

    print(table)

with create_app().app_context():
    customer = Customer.query.filter_by(name='US Sugar').first()

    if not customer:
        print('Customer not found!')
        exit()

    sensors = Sensor.query.filter_by(customer_id=customer.id).all()
    result = [sensor.as_dict() for sensor in sensors]
    df = pd.DataFrame(result)

    # Sort the DataFrame by 'source_status' and then by 'device_model'
    sorted_df = df.sort_values(by=['source_status', 'source_name'], ascending=[False,True])



    filename = "sorted_data.xlsx"
    writer = pd.ExcelWriter(filename, engine='openpyxl')

    sorted_df.to_excel(writer, sheet_name='Sheet1', index=False)

    workbook = writer.book
    worksheet = writer.sheets['Sheet1']

    red_fill = openpyxl.styles.PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    green_fill = openpyxl.styles.PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    amber_fill = openpyxl.styles.PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')


    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):  # starting from 2 to skip the header
        source_status_cell = row[4]  # Adjust the index (2) if 'source_status' is not the third column
        if source_status_cell.value == 'UNRESPONSIVE':
            source_status_cell.fill = red_fill
        if source_status_cell.value == 'RESPONSIVE':
            source_status_cell.fill = green_fill

        source_name_cell = row[1]
        battery_voltage_cell = row[5]  # Adjust the index (2) if 'source_status' is not the third column

        if int(battery_voltage_cell.value) < 2700 and source_name_cell.value == "MaxActive - Cellular":
            print('active')
            battery_voltage_cell.fill = red_fill
        if int(battery_voltage_cell.value) < 10000 and source_name_cell.value == "MaxTracker":
            print('tracker')
            battery_voltage_cell.fill = red_fill

        battery_status_cell = row[6]
        if battery_status_cell.value == "DEAD" or battery_status_cell.value == "DISCONNECTED":
            battery_status_cell.fill = red_fill

    writer.close()
